import os

from openpyxl import Workbook

from GenerationAwareAnalyze import record


def summarize_generation_aware_test_plan(project_root: str):
    workbook = Workbook()

    sheet_name = 'Generation-aware Tests Summarization'
    if sheet_name in workbook.sheetnames:
        workbook.remove(sheet_name)
    workbook.create_sheet(sheet_name)
    
    generation_aware_sheet = workbook[sheet_name]

    for project_name in os.listdir(project_root):
        project_path = os.path.join(project_root, project_name)

        # if it's not a folder, just ignore
        if not os.path.isdir(project_path):
            continue

        # collect week info
        week_info_path = os.path.join(project_path, 'week.json')
        week_info = record.load_week_info_from_json(week_info_path)
        monday_date = week_info['monday'].replace('-', '/')
        friday_date = week_info['friday'].replace('-', '/')

        # collect generation aware test info
        generation_aware_info_path = os.path.join(project_path, 'GenerationAwareAnalysis.json')
        generation_aware_info = record.load_generation_aware_test_info_to_json(generation_aware_info_path)
        runtime_commit = generation_aware_info.runtime_commit_info['sha']

        row = [
            f'{monday_date} ~ {friday_date}',
            runtime_commit,
        ]

        generation_aware_sheet.append(row)

    summarization_path = os.path.join(project_root, 'GenerationAwareAnalysisSum.xlsx')
    if os.path.exists(summarization_path):
        os.remove(summarization_path)
    workbook.save(summarization_path)