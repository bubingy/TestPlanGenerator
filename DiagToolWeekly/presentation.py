from openpyxl import Workbook
from openpyxl.styles import Font

import app
from DiagToolWeekly.dotnet_sdk import DotnetSDKInfo
from DiagToolWeekly.diag_tools import DiagnosticToolsInfo
from DiagToolWeekly.configuration import LTTngWeeklyTestconfig


@app.function_monitor(
    pre_run_msg='write week info to spread sheet'
)
def write_week_info_to_sheet(workbook: Workbook, week_info: dict) -> None:
    sheet_name = 'week info'
    if sheet_name in workbook.sheetnames:
        workbook.remove(sheet_name)
    workbook.create_sheet(sheet_name)
    
    week_info_sheet = workbook[sheet_name]
    
    monday = '/'.join(week_info['monday'].split('-'))
    friday = '/'.join(week_info['friday'].split('-'))
    print(f'Week({monday} ~ {friday})')
    week_info_sheet.cell(
        row=1, column=1,
        value=f'Week({monday} ~ {friday})'
    )


@app.function_monitor(
    pre_run_msg='write sdk info to spread sheet'
)
def write_sdk_info_list_to_sheet(workbook: Workbook, sdk_info_list: list[DotnetSDKInfo]) -> None:
    sheet_name = 'SDKVersion'
    if sheet_name in workbook.sheetnames:
        workbook.remove(sheet_name)
    workbook.create_sheet(sheet_name)
    
    sdk_version_sheet = workbook[sheet_name]
    print('Full version of sdk')
    sdk_version_sheet.cell(
        row=1, column=1,
        value='Branch of sdk'
    )
    sdk_version_sheet.cell(
        row=1, column=2,
        value='Version of sdk'
    )
    sdk_version_sheet.cell(
        row=1, column=3,
        value='Build ID of sdk'
    )

    for idx, sdk_info in enumerate(sdk_info_list):
        print(f'{sdk_info.branch_name}: {sdk_info.dotnet_sdk_version} build id: {sdk_info.build_id}')
        sdk_version_sheet.cell(
            row=2+idx, column=1,
            value=f'{sdk_info.branch_name}:'
        )
        sdk_version_sheet.cell(
            row=2+idx, column=2,
            value=f'{sdk_info.dotnet_sdk_version}'
        )
        sdk_version_sheet.cell(
            row=2+idx, column=3,
            value=f'{sdk_info.build_id}'
        )


@app.function_monitor(
    pre_run_msg='write diag tool info to spread sheet'
)
def write_diag_tools_info_to_sheet(workbook: Workbook, diag_tools_info: DiagnosticToolsInfo) -> None:
    sheet_name = 'ToolInfo'
    if sheet_name in workbook.sheetnames:
        workbook.remove(sheet_name)
    workbook.create_sheet(sheet_name)

    tool_info_sheet = workbook[sheet_name]
    print('Info of tool:')
    tool_info_sheet.cell(
        row=1, column=1,
        value='Info of tool'
    )
    print(f'Version: {diag_tools_info.diag_tool_version}')
    tool_info_sheet.cell(
        row=2, column=1,
        value=f'Version'
    )
    tool_info_sheet.cell(
        row=2, column=2,
        value=f'{diag_tools_info.diag_tool_version}'
    )
    tool_info_sheet.cell(
        row=3, column=1,
        value='Build ID'
    )
    tool_info_sheet.cell(
        row=3, column=2,
        value=f'{diag_tools_info.build_id}'
    )


def init_sos_unit_content(os_name: str, sys_ptrace_enabled=False):
    if 'alpine' in os_name.lower():
        return 'NA'
    if sys_ptrace_enabled is True:
        return 'NA'
    return ''


def init_dump_unit_content(os_name: str, sdk_version: str):
    if 'osx' in os_name.lower():
        if sdk_version[:3] == '6.0':
            return 'NA'
        else:
            return ''
    else:
        return ''


@app.function_monitor(
    pre_run_msg='write diag tool test matrix to spread sheet'
)
def write_weekly_diag_tools_test_matrix_to_sheet(workbook: Workbook, os_sdk_map: dict) -> None:
    sheet_name = 'Diagnostic Tools Test'
    if sheet_name in workbook.sheetnames:
        workbook.remove(sheet_name)
    workbook.create_sheet(sheet_name)

    test_matrix_sheet = workbook[sheet_name]

    header = [
        '',
        'dotnet-counters',
        'dotnet-dump',
        'dotnet-gcdump',
        'dotnet-sos',
        'dotnet-stack',
        'dotnet-trace'
    ]
    for idx, title in enumerate(header):
        test_matrix_sheet.cell(
            row=1, column=1+idx,
            value=title
        ).font = Font(color="1F497D")

    # write required os
    required_oses = os_sdk_map['requiredOS']
    current_row = 2
    for idx, key in enumerate(required_oses.keys()):
        os_name = key
        sdk_version = required_oses[key]

        if 'Alpine' in key:
            test_matrix_sheet.cell(
                row=current_row+idx, column=1,
                value=f'{os_name}/{sdk_version}'
            ).font = Font(color="1F497D")
            test_matrix_sheet.cell(
                row=current_row+idx, column=3,
                value=init_dump_unit_content(os_name, sdk_version)
            ).font = Font(color="5B9BD5")
            test_matrix_sheet.cell(
                row=current_row+idx, column=5,
                value=init_sos_unit_content(os_name, False)
            ).font = Font(color="5B9BD5")
            current_row += 1

            test_matrix_sheet.cell(
                row=current_row+idx, column=1,
                value=f'{os_name}/{sdk_version} disable SYS_PTACE, seccomp=unconfined'
            ).font = Font(color="1F497D")
            test_matrix_sheet.cell(
                row=current_row+idx, column=3,
                value=init_dump_unit_content(os_name, sdk_version)
            ).font = Font(color="5B9BD5")
            test_matrix_sheet.cell(
                row=current_row+idx, column=5,
                value=init_sos_unit_content(os_name, False)
            ).font = Font(color="5B9BD5")
            continue

        test_matrix_sheet.cell(
            row=current_row+idx, column=1,
            value=f'{os_name}/{sdk_version}'
        ).font = Font(color="1F497D")
        test_matrix_sheet.cell(
            row=current_row+idx, column=3,
            value=init_dump_unit_content(os_name, sdk_version)
        ).font = Font(color="5B9BD5")
        test_matrix_sheet.cell(
            row=current_row+idx, column=5,
            value=init_sos_unit_content(os_name, False)
        ).font = Font(color="5B9BD5")
    current_row = current_row + idx + 1

    # write alternate os
    alternate_oses = os_sdk_map['alternateOS'] 
    for idx, key in enumerate(alternate_oses.keys()):
        os_name = alternate_oses[key]
        sdk_version = key

        test_matrix_sheet.cell(
            row=current_row+idx, column=1,
            value=f'{os_name}/{sdk_version}'
        ).font = Font(color="1F497D")
        test_matrix_sheet.cell(
            row=current_row+idx, column=3,
            value=init_dump_unit_content(os_name, sdk_version)
        ).font = Font(color="5B9BD5")
        test_matrix_sheet.cell(
            row=current_row+idx, column=5,
            value=init_sos_unit_content(os_name, False)
        ).font = Font(color="5B9BD5")
        current_row += 1

        test_matrix_sheet.cell(
            row=current_row+idx, column=1,
            value=f'{os_name}/{sdk_version} disable SYS_PTACE, seccomp=unconfined'
        ).font = Font(color="1F497D")
        test_matrix_sheet.cell(
            row=current_row+idx, column=3,
            value=init_dump_unit_content(os_name, sdk_version)
        ).font = Font(color="5B9BD5")
        test_matrix_sheet.cell(
            row=current_row+idx, column=5,
            value=init_sos_unit_content(os_name, True)
        ).font = Font(color="5B9BD5")


@app.function_monitor(
    pre_run_msg='write lttng test matrix to spread sheet'
)
def print_weekly_lttng_test_matrix_on_sheet(workbook: Workbook, lttng_conf: LTTngWeeklyTestconfig) -> None:
    sheet_name = 'LTTng Test'
    if sheet_name in workbook.sheetnames:
        workbook.remove(sheet_name)
    workbook.create_sheet(sheet_name)
    test_matrix_sheet = workbook[sheet_name]

    header = ['']
    for os_name in lttng_conf.lttng_os_list:
        header.append(os_name)
    for idx, title in enumerate(header):
        test_matrix_sheet.cell(
            row=1, column=1+idx,
            value=title
        ).font = Font(color="1F497D")

    for idx, sdk_major_version in enumerate(lttng_conf.SDK_major_version_list):
        test_matrix_sheet.cell(
            row=2+idx, column=1,
            value=sdk_major_version
        ).font = Font(color="1F497D")
