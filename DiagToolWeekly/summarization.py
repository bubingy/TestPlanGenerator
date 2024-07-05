import os

from openpyxl import Workbook

from DiagToolWeekly import record


def summarize_diag_tool_test_plan(project_root: str):
    workbook = Workbook()

    # diag tool test summarization
    sheet_name = 'Diag Tool Tests Summarization'
    if sheet_name in workbook.sheetnames:
        workbook.remove(sheet_name)
    workbook.create_sheet(sheet_name)
    
    diag_tool_sheet = workbook[sheet_name]

    for project_name in os.listdir(project_root):
        project_path = os.path.join(project_root, project_name)

        # if it's not a folder, just ignore
        if not os.path.isdir(project_path):
            continue

        # collect week info
        week_info_path = os.path.join(project_path, 'week.json')
        week_info = record.load_week_info_from_json(week_info_path)
        monday_date = week_info['monday'].replace('-', '/')
        friday_date = week_info['friday'].replace('-', '/')

        # collect diag tool info
        diag_tool_info_path = os.path.join(project_path, 'DiagToolInfo.json')
        diag_tool_info = record.load_diag_tool_info_from_json(diag_tool_info_path)
        tool_version = diag_tool_info['version']

        # collect os & sdk map
        diag_tool_test_matrix_path = os.path.join(project_path, 'DiagToolTestMatrix.json')
        diag_tool_test_matrix = record.load_diag_tool_test_matrix_from_json(diag_tool_test_matrix_path)

        row = [
            f'{monday_date} ~ {friday_date}',
            tool_version,
        ]

        for os_name in diag_tool_test_matrix['requiredOS'].keys():
            branch = diag_tool_test_matrix['requiredOS'][os_name]
            row.append(f'{os_name}/{branch}')

        for branch in diag_tool_test_matrix['alternateOS'].keys():
            os_name = diag_tool_test_matrix['alternateOS'][branch]
            row.append(f'{os_name}/{branch}')

        diag_tool_sheet.append(row)

    # LTTng test summarization
    sheet_name = 'LTTng Tests Summarization'
    if sheet_name in workbook.sheetnames:
        workbook.remove(sheet_name)
    workbook.create_sheet(sheet_name)
    
    lttng_test_sheet = workbook[sheet_name]

    for project_name in os.listdir(project_root):
        project_path = os.path.join(project_root, project_name)

        # if it's not a folder, just ignore
        if not os.path.isdir(project_path):
            continue
        
        # collect week info
        week_info_path = os.path.join(project_path, 'week.json')
        week_info = record.load_week_info_from_json(week_info_path)
        monday_date = week_info['monday'].replace('-', '/')
        friday_date = week_info['friday'].replace('-', '/')

        sdk_info_path = os.path.join(project_path, 'SDKInfo.json')
        sdk_info_list = record.load_SDK_info_from_json(sdk_info_path)

        row = [
            f'{monday_date} ~ {friday_date}'
        ]
        for sdk_info in sdk_info_list:
            row.append(sdk_info.dotnet_sdk_version)

        lttng_test_sheet.append(row)

    summarization_path = os.path.join(project_root, 'summarization.xlsx')
    if os.path.exists(summarization_path):
        os.remove(summarization_path)
    workbook.save(summarization_path)